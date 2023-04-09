#!/usr/bin/env python
# coding: utf-8

# 
# # Project #2 -- Academic Cups
#     
# 
# ## Each year, the Lawrenceville School awards two academic cups to Houses:
# <blockquote>
#     
# - ### Green Cup:  Awarded to the Circle or Crescent House who has the largest GPA increase between the fall and spring terms
# <br>
#     
# - ### Chivers Cup:  Awarded to the Circle or Crescent House who has the highest cumulative GPA
#  
# </blockquote>
# 
# ## Your task is to write a program that finds the top three Houses in each category.  Here are the details:
# <blockquote>
#     
# ### 1. There is a file called Grades_2021.xlsx that is located in your "Projects" folder that contains all the grade data from the 2021 school year.  The data is real, exported directly out of Veracross, but all student names have been removed.  Note that Veracross stores all letter grades as two characters (for +'s and -'s), so grades will have a space after them if there isn't a + or -.
# 
# ### 2. Only Circle & Crescent Houses compete in these two cups, and only grades earned by 3rd and 4th formers count towards the House GPAs (no prefects).  So you will need to do some cleaning of the data.
#     
# ### 3. You are not permitted to alter the Excel file in any way... the data must be read as is and handled in your program.  However, if you wish to use Python to create a new cleaned-up file, you are perfectly welcome to do so.
#     
# ### 4. When you present the top Houses in each category, you should list the average GPA for each House to 4 decimal places.
# 
# ### 5. This program should be usable, so your final product should allow the user change the file being accessed.
# </blockquote>

# In[1]:


## to Mr. F: less efficient version in this cell; more efficient version (using functions) in next cell 
from openpyxl import load_workbook
from openpyxl import Workbook
wb = load_workbook(filename = "Grades_2021.xlsx")  # Loads the file
sheet = wb['Sheet1']      
def letter_to_num(x):
    my_dict = {'A+': 4.3, 'A ': 4.0, 'A-': 3.7, 'B+': 3.3, 'B ': 3.0, 'B-': 2.7, 'C+': 2.3, 'C ': 2.0, 'C-': 1.7, 
               'D+': 1.3, 'D ': 1.0, 'D-': 0.7}
    return(my_dict[x]) ## create function, letter_to_num that converts between letter grade and corresponding num GPA using dictionaries (keys and values)
for i in range(2,sheet.max_row + 1):
    house = sheet.cell(row = i,column = 1).value
    grade_level = sheet.cell(row = i, column = 2).value
    if grade_level == "V" or grade_level == "II" or grade_level == "PG": # cleaning the data to only III/IV formers
        continue
    letter_grade = sheet.cell(row = i, column = 4).value   

houses_dict = {'Carter': [], 'Cleve': [],'Dickinson': [], 'Griswold': [], 'Hamill': [], 'Kennedy': [], 
               'Kirby': [], 'McClellan': [], 'Stanley': [],
              'Stephens': [], 'Woodhull': []} # create dictionary with empty values so that we can fill them with
## the grades for corresponding house

for key,value in houses_dict.items():
    for i in range(2, sheet.max_row + 1):      
        if key == (sheet.cell(row = i,column = 1).value):
            houses_dict[key] += [letter_to_num(sheet.cell(row = i, column = 4).value)]
my_list = []
for key, value in houses_dict.items():
    house_GPA = round(sum(houses_dict[key])/len(houses_dict[key]), 4)
#     print(f'The average GPA for {key} is {house_GPA}.')
    my_list.append(house_GPA)
# print(my_list)
houses_list = ['Carter', 'Cleve','Dickinson', 'Griswold', 'Hamill', 'Kennedy', 
               'Kirby', 'McClellan', 'Stanley',
              'Stephens', 'Woodhull']

top_three = sorted(zip(my_list, houses_list), reverse=True)[:3] ## zips lists, sorts them in reverse order, then takes first three (ie. the houses with the top three GPAs)
top_three_dict = dict(top_three) ## converts data type from list to dictionary 

print('The top three houses for the Chivers Cup are:')
for key, value in top_three_dict.items(): ## iterate over each key and value within the houses with top 3 GPAs
        print(f' {value} with an average GPA of {key}.')
## after this f string, the code for Chiver's is complete. Now we can move onto Green cup code: 
## ----------------------------------------------------------------------------------------------------------------
        
t1_houses_dict = {'Carter': [], 'Cleve': [],'Dickinson': [], 'Griswold': [], 'Hamill': [], 'Kennedy': [], 
               'Kirby': [], 'McClellan': [], 'Stanley': [],
              'Stephens': [], 'Woodhull': []}

for key,value in t1_houses_dict.items():
    for i in range(2, sheet.max_row + 1):  
        term = sheet.cell(row = i, column = 3).value
        grade_level = sheet.cell(row = i, column = 2).value 
        if term == "T2" or term == "T3" or grade_level == "V" or grade_level == "II" or grade_level == "PG":
            continue # skip over term 2, term 3, second formers/fifth formers/PGS for data cleaning/getting term 1 grades
            #it is important to get term 1 grades because green cup relies on the difference between fall and spring GPAs
        if key == (sheet.cell(row = i,column = 1).value):
            t1_houses_dict[key] += [letter_to_num(sheet.cell(row = i, column = 4).value)]
                                
t1_my_list = []
for key, value in t1_houses_dict.items():
    t1_house_GPA = round(sum(t1_houses_dict[key])/len(t1_houses_dict[key]), 4)
#     print(f'The average GPA for {key} is {t1_house_GPA}.')
    t1_my_list.append(t1_house_GPA)
# print(t1_my_list)

houses_list = ['Carter', 'Cleve','Dickinson', 'Griswold', 'Hamill', 'Kennedy', 
               'Kirby', 'McClellan', 'Stanley',
              'Stephens', 'Woodhull']

top_three_t1 = sorted(zip(houses_list, t1_my_list)) ## zips the two lists to get houses and corresponding GPAs
top_three_dict_t1 = dict(top_three_t1)
# print(top_three_dict_t1) ## dictionary with house as key and term 1 GPA (numerical) as associated value

## same approach as above, but for term 3:
t3_houses_dict = {'Carter': [], 'Cleve': [],'Dickinson': [], 'Griswold': [], 'Hamill': [], 'Kennedy': [], 
               'Kirby': [], 'McClellan': [], 'Stanley': [],
              'Stephens': [], 'Woodhull': []}

for key,value in t3_houses_dict.items():
    for i in range(2, sheet.max_row + 1):  
        term = sheet.cell(row = i, column = 3).value
        grade_level = sheet.cell(row = i, column = 2).value 
        if term == "T1" or term == "T2" or grade_level == "V" or grade_level == "II" or grade_level == "PG":
            continue
        if key == (sheet.cell(row = i,column = 1).value):
            t3_houses_dict[key] += [letter_to_num(sheet.cell(row = i, column = 4).value)]
t3_my_list = []
for key, value in t3_houses_dict.items():
    t3_house_GPA = round(sum(t3_houses_dict[key])/len(t3_houses_dict[key]), 4)
#     print(f'The average GPA for {key} is {t3_house_GPA}.')
    t3_my_list.append(t3_house_GPA) ## add the third term GPA to each house
# print(t3_house_GPA)
# print(t3_my_list)

houses_list = ['Carter', 'Cleve','Dickinson', 'Griswold', 'Hamill', 'Kennedy', 
               'Kirby', 'McClellan', 'Stanley',
              'Stephens', 'Woodhull']

top_three_t3 = sorted(zip(houses_list, t3_my_list))
top_three_dict_t3 = dict(top_three_t3)
# print(top_three_dict_t3) ## dictionary with house as key and term 3 GPA (numerical) as associated value


## now that we have terms 1 and terms 3, we can find the difference to see the top 3 houses for Green Cup!
## let's use some dictionary comprehension!!
t1_t3_diff = {key: round(top_three_dict_t3[key] - top_three_dict_t1.get(key, 0), 4) ## round to 4 decimal points
                       for key in top_three_dict_t3.keys()} ## got method for subtracting elements in two
# dictionaries from: https://www.geeksforgeeks.org/python-subtraction-of-dictionaries/
## essentially, what the above is doing is using dictionary comprehension (which we learned) 
#+ keys(), a shorthand to the longer method of loops to subtract each value of t1 from each value of t3
## This is important because to calculate Green cup, you need the difference in term GPAs between spring and fall!
# print(t1_t3_diff)

diff_list = [] 
for key, value in t1_t3_diff.items():
    diff_list.append(value) ## create a list that has the t3-t1 differences
# print(diff_list) ## list with all t3-t1 differences as elements 

diff_dict = dict(sorted(zip(diff_list, houses_list), reverse=True)[:3]) ## zip the differences list to the house list
# print(diff_dict) 

print('The top three houses for the Green Cup are:')
for key, value in diff_dict.items():
    print(f' {value} with a fall-to-spring GPA increase of {key:.4f}.') # f string to print house and increase in house GPA
    # for top three Green cuppers


# ## Once you have the Chivers & Green Cup portions of the program written, you can work on any of the following to continue demonstrating your proficiency in the various standards.   I would strongly suggest making a copy of your working program before you start adding stuff!
# <blockquote>
#     
# ### 1. Build out your interface to allow a user to print out the different grades for each House.
#    
# ### 2. Build a statistics function to allow the user to find things like mean, median, standard deviation, etc. for all grades and/or for individual Houses.
#     
# ### 3. Anything else you can think of!
#     
# </blockquote>

# In[2]:


## as noted above, this cell contains the more efficient version, using functions!
from openpyxl import load_workbook
from openpyxl import Workbook
wb = load_workbook(filename = "Grades_2021.xlsx")  # Loads the file
sheet = wb['Sheet1']      
def letter_to_num(x):
    my_dict = {'A+': 4.3, 'A ': 4.0, 'A-': 3.7, 'B+': 3.3, 'B ': 3.0, 'B-': 2.7, 'C+': 2.3, 'C ': 2.0, 'C-': 1.7, 
               'D+': 1.3, 'D ': 1.0, 'D-': 0.7}
    return(my_dict[x]) # creating same function that converts letter grades to numerical grades

for i in range(2,sheet.max_row + 1):
    house = sheet.cell(row = i,column = 1).value
    grade_level = sheet.cell(row = i, column = 2).value
    if grade_level == "V" or grade_level == "II" or grade_level == "PG": # cleaning the data to only III/IV formers
        continue
    letter_grade = sheet.cell(row = i, column = 4).value   

houses_dict = {'Carter': [], 'Cleve': [],'Dickinson': [], 'Griswold': [], 'Hamill': [], 'Kennedy': [], 
               'Kirby': [], 'McClellan': [], 'Stanley': [],
              'Stephens': [], 'Woodhull': []}

for key,value in houses_dict.items():
    for i in range(2, sheet.max_row + 1):      
        if key == (sheet.cell(row = i,column = 1).value):
            houses_dict[key] += [letter_to_num(sheet.cell(row = i, column = 4).value)]
my_list = []
for key, value in houses_dict.items():
    house_GPA = round(sum(houses_dict[key])/len(houses_dict[key]), 4)
#     print(f'The average GPA for {key} is {house_GPA}.')
    my_list.append(house_GPA)

houses_list = ['Carter', 'Cleve','Dickinson', 'Griswold', 'Hamill', 'Kennedy', 
               'Kirby', 'McClellan', 'Stanley',
              'Stephens', 'Woodhull']

top_three = sorted(zip(my_list, houses_list), reverse=True)[:3]
top_three_dict = dict(top_three)

print('The top three houses for the Chivers Cup are:')
for key, value in top_three_dict.items():
        print(f' {value} with an average GPA of {key}.')
        
### CHIVERS DONE (same methods as above; effiency differences come with Green cup) -----------------------------------------------------------------------------------------------
houses_dict = {'Carter': [], 'Cleve': [],'Dickinson': [], 'Griswold': [], 'Hamill': [], 'Kennedy': [], 
               'Kirby': [], 'McClellan': [], 'Stanley': [],
              'Stephens': [], 'Woodhull': []}
def term_GPA_func(tx): # define function, term_GPA_func, with one argument, tx
    for key,value in houses_dict.items():
        for i in range(2, sheet.max_row + 1):  
            term = sheet.cell(row = i, column = 3).value
            grade_level = sheet.cell(row = i, column = 2).value 
            if term != tx or grade_level == "V" or grade_level == "II" or grade_level == "PG":
                continue # clean data to not include freshmen, seniors, and PGs; 
                # gets the GPA of a specefied term by using "continue" when the term != tx
            if key == (sheet.cell(row = i,column = 1).value):
                houses_dict[key] += [letter_to_num(sheet.cell(row = i, column = 4).value)]
    return houses_dict
# print(term_GPA_func('T1'))
# term_GPA_func("T1") yeilds a dict with all grades for each house as key and grades (nums) as values
# # term_GPA_func("T3") the below works! ------------------------------------------------------------------------

t1_grades = []
for key, value in term_GPA_func('T1').items():
    t1_house_GPA = sum(houses_dict[key])/len(houses_dict[key])
    t1_grades.append(t1_house_GPA)
#     print(f'The average GPA of {key} for term 1 was {t1_house_GPA}.')
# print(t1_grades)

houses_dict = {'Carter': [], 'Cleve': [],'Dickinson': [], 'Griswold': [], 'Hamill': [], 'Kennedy': [], 
               'Kirby': [], 'McClellan': [], 'Stanley': [],
              'Stephens': [], 'Woodhull': []}

t3_grades = []
for key, value in term_GPA_func('T3').items():
    t3_house_GPA = sum(houses_dict[key])/len(houses_dict[key])
    t3_grades.append(t3_house_GPA)
#     print(f'The average GPA of {key} for term 3 was {t3_house_GPA}.')
# print(t3_grades)

diff = [] ## initializing empty array
for i, j in zip(t1_grades, t3_grades):
    q = round((j-i), 4)
    diff.append(q) ## adding each differences to initially empty list, called diff
# print(diff)

diff_dict = dict(sorted(zip(diff, houses_list), reverse=True)[:3]) ## zip diff and houses_list to get difference with correspinding house
## sort this zipped list in reverse so order is highest to lowest GPAs, then take first three houses (ie. the three with highest increase in GPA)
# print(diff_dict)

print('The top three houses for the Green Cup are:')
for key, value in diff_dict.items():
    print(f' {value} with a fall-to-spring GPA increase of {key:.4f}.') ## round to 4 decimal points, as instructions outlined!


# In[ ]:




